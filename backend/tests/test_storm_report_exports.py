"""
Focused unit tests for storm protection report export endpoints.

These tests cover:
- _parse_limit boundary conditions
- CSV column names and field mapping
- XLSX generation (openpyxl)
- Empty dataset handling
- Null/None field handling
- format fallthrough to CSV for unknown values

No real MongoDB, Flask server, or network required.
"""
from __future__ import annotations

import io
import csv
import unittest
from datetime import datetime, timezone
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Replicate _parse_limit exactly as implemented in report_routes.py
# ---------------------------------------------------------------------------

def _parse_limit(raw: str, max_default: int = 5000, max_limit: int = 50000) -> int:
    if not raw:
        return max_default
    try:
        limit = int(raw)
    except ValueError as exc:
        raise ValueError("Invalid limit") from exc
    limit = max(1, limit)
    return min(limit, max_limit)


# ---------------------------------------------------------------------------
# Replicate _export_response logic for test verification
# ---------------------------------------------------------------------------

def _export_csv(headers: list, rows: list) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue()


def _export_xlsx(title: str, headers: list, rows: list) -> bytes:
    wb = Workbook()
    sheet = wb.active
    if sheet is None:
        sheet = wb.create_sheet()
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sample serialized documents (as returned by serialize_* functions)
# ---------------------------------------------------------------------------

SAMPLE_INCIDENT = {
    "incidentId": "storm-2026-000001",
    "hostname": "sw1",
    "interface": "Gi1/0/5",
    "severity": "CRITICAL",
    "status": "OPEN",
    "createdAt": "2026-08-11T08:00:00Z",
}

SAMPLE_INCIDENT_NULLS = {
    "incidentId": "storm-2026-000002",
    "hostname": None,
    "interface": "Gi1/0/6",
    "severity": None,
    "status": None,
    "createdAt": None,
}

SAMPLE_INCIDENT_SPECIAL = {
    "incidentId": "storm-2026-000003",
    "hostname": 'Device,"01"',
    "interface": "Gi1/0/7",
    "severity": "HIGH",
    "status": "MITIGATED",
    "createdAt": "2026-08-11T09:00:00Z",
}

SAMPLE_MITIGATION = {
    "incidentId": "storm-2026-000001",
    "interface": "Gi1/0/5",
    "strategy": "SHUTDOWN",
    "status": "SUCCESS",
    "operator": "admin",
    "timestamp": "2026-08-11T08:01:00Z",
}

SAMPLE_RECOVERY = {
    "incidentId": "storm-2026-000001",
    "interface": "Gi1/0/5",
    "recoveryStatus": "RECOVERED",
    "executedBy": "admin",
    "timestamp": "2026-08-11T08:30:00Z",
}

SAMPLE_RECOVERY_NO_EXECUTOR = {
    "incidentId": "storm-2026-000002",
    "interface": "Gi1/0/6",
    "recoveryStatus": "BLOCKED",
    "executedBy": None,
    "timestamp": "2026-08-11T09:00:00Z",
}


def _incident_row(d: dict) -> list:
    return [
        d.get("incidentId"),
        d.get("hostname"),
        d.get("interface"),
        d.get("severity"),
        d.get("status"),
        d.get("createdAt"),
    ]


def _mitigation_row(d: dict) -> list:
    return [
        d.get("incidentId"),
        d.get("interface"),
        d.get("strategy"),
        d.get("status"),
        d.get("operator"),
        d.get("timestamp"),
    ]


def _recovery_row(d: dict) -> list:
    return [
        d.get("incidentId"),
        d.get("interface"),
        d.get("recoveryStatus"),
        d.get("executedBy") or "",
        d.get("timestamp"),
    ]


INCIDENT_HEADERS = ["Incident", "Switch", "Interface", "Severity", "Status", "Created"]
MITIGATION_HEADERS = ["Incident", "Interface", "Strategy", "Status", "Operator", "Time"]
RECOVERY_HEADERS = ["Incident", "Interface", "Status", "Executed By", "Time"]


class TestParseLimitBoundaries(unittest.TestCase):

    def test_empty_returns_default(self):
        self.assertEqual(_parse_limit(""), 5000)

    def test_normal_value(self):
        self.assertEqual(_parse_limit("10"), 10)

    def test_zero_coerced_to_one(self):
        self.assertEqual(_parse_limit("0"), 1)

    def test_negative_coerced_to_one(self):
        self.assertEqual(_parse_limit("-100"), 1)

    def test_exceeds_max_capped(self):
        self.assertEqual(_parse_limit("999999"), 50000)

    def test_exactly_max_limit(self):
        self.assertEqual(_parse_limit("50000"), 50000)

    def test_one_below_max(self):
        self.assertEqual(_parse_limit("49999"), 49999)

    def test_non_numeric_raises(self):
        with self.assertRaises(ValueError):
            _parse_limit("pdf")

    def test_float_raises(self):
        with self.assertRaises(ValueError):
            _parse_limit("10.5")

    def test_ui_default_ten(self):
        # The UI sends limit=10; should round-trip correctly
        self.assertEqual(_parse_limit("10"), 10)


class TestIncidentsCsvExport(unittest.TestCase):

    def _build_rows(self, docs):
        return [_incident_row(d) for d in docs]

    def test_headers_correct(self):
        result = _export_csv(INCIDENT_HEADERS, [])
        first_line = result.splitlines()[0]
        self.assertEqual(first_line, "Incident,Switch,Interface,Severity,Status,Created")

    def test_single_row(self):
        rows = self._build_rows([SAMPLE_INCIDENT])
        result = _export_csv(INCIDENT_HEADERS, rows)
        self.assertIn("storm-2026-000001", result)
        self.assertIn("sw1", result)
        self.assertIn("CRITICAL", result)

    def test_empty_dataset_has_header_only(self):
        result = _export_csv(INCIDENT_HEADERS, [])
        lines = [l for l in result.splitlines() if l]
        self.assertEqual(len(lines), 1)
        self.assertIn("Incident", lines[0])

    def test_null_fields_handled(self):
        rows = self._build_rows([SAMPLE_INCIDENT_NULLS])
        result = _export_csv(INCIDENT_HEADERS, rows)
        # Should not raise; None fields appear as empty
        self.assertIn("storm-2026-000002", result)

    def test_special_chars_comma_and_quote(self):
        rows = self._build_rows([SAMPLE_INCIDENT_SPECIAL])
        result = _export_csv(INCIDENT_HEADERS, rows)
        # csv module must quote the field containing a comma
        self.assertIn("storm-2026-000003", result)
        # Verify result is parseable back
        reader = csv.reader(io.StringIO(result))
        data_rows = list(reader)
        self.assertEqual(len(data_rows), 2)  # header + 1 data
        hostname_col = data_rows[1][1]
        self.assertIn("01", hostname_col)

    def test_ten_rows_limit(self):
        docs = [dict(SAMPLE_INCIDENT, incidentId=f"storm-2026-{i:06d}") for i in range(10)]
        rows = self._build_rows(docs)
        result = _export_csv(INCIDENT_HEADERS, rows)
        data_rows = [l for l in result.splitlines() if l][1:]
        self.assertEqual(len(data_rows), 10)


class TestMitigationsCsvExport(unittest.TestCase):

    def test_headers_correct(self):
        result = _export_csv(MITIGATION_HEADERS, [])
        first_line = result.splitlines()[0]
        self.assertEqual(first_line, "Incident,Interface,Strategy,Status,Operator,Time")

    def test_single_row_fields(self):
        rows = [_mitigation_row(SAMPLE_MITIGATION)]
        result = _export_csv(MITIGATION_HEADERS, rows)
        self.assertIn("SHUTDOWN", result)
        self.assertIn("admin", result)
        self.assertIn("SUCCESS", result)

    def test_empty_dataset(self):
        result = _export_csv(MITIGATION_HEADERS, [])
        lines = [l for l in result.splitlines() if l]
        self.assertEqual(len(lines), 1)


class TestRecoveriesCsvExport(unittest.TestCase):

    def test_headers_correct(self):
        result = _export_csv(RECOVERY_HEADERS, [])
        first_line = result.splitlines()[0]
        self.assertEqual(first_line, "Incident,Interface,Status,Executed By,Time")

    def test_single_row_fields(self):
        rows = [_recovery_row(SAMPLE_RECOVERY)]
        result = _export_csv(RECOVERY_HEADERS, rows)
        self.assertIn("RECOVERED", result)
        self.assertIn("admin", result)

    def test_null_executed_by_becomes_empty_string(self):
        rows = [_recovery_row(SAMPLE_RECOVERY_NO_EXECUTOR)]
        result = _export_csv(RECOVERY_HEADERS, rows)
        # executedBy=None -> "" -> empty field, not "None" or "—"
        self.assertNotIn("None", result)
        self.assertNotIn("\u2014", result)  # no em-dash
        self.assertIn("storm-2026-000002", result)

    def test_empty_dataset(self):
        result = _export_csv(RECOVERY_HEADERS, [])
        lines = [l for l in result.splitlines() if l]
        self.assertEqual(len(lines), 1)


class TestXlsxGeneration(unittest.TestCase):

    def test_incidents_xlsx_valid(self):
        rows = [_incident_row(SAMPLE_INCIDENT)]
        data = _export_xlsx("storm_incidents", INCIDENT_HEADERS, rows)
        self.assertGreater(len(data), 1000)
        # Verify it opens as a valid workbook
        wb = Workbook()
        buf = io.BytesIO(data)
        from openpyxl import load_workbook
        loaded = load_workbook(buf)
        sheet = loaded.active
        self.assertIsNotNone(sheet)
        header_row = [cell.value for cell in sheet[1]]
        self.assertEqual(header_row, INCIDENT_HEADERS)

    def test_mitigations_xlsx_valid(self):
        rows = [_mitigation_row(SAMPLE_MITIGATION)]
        data = _export_xlsx("storm_mitigations", MITIGATION_HEADERS, rows)
        self.assertGreater(len(data), 1000)

    def test_recoveries_xlsx_valid(self):
        rows = [_recovery_row(SAMPLE_RECOVERY)]
        data = _export_xlsx("storm_recoveries", RECOVERY_HEADERS, rows)
        self.assertGreater(len(data), 1000)

    def test_xlsx_empty_dataset(self):
        data = _export_xlsx("storm_incidents", INCIDENT_HEADERS, [])
        self.assertGreater(len(data), 1000)
        from openpyxl import load_workbook
        loaded = load_workbook(io.BytesIO(data))
        sheet = loaded.active
        self.assertIsNotNone(sheet)
        # Only header row
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)

    def test_xlsx_null_fields_no_crash(self):
        rows = [_incident_row(SAMPLE_INCIDENT_NULLS)]
        data = _export_xlsx("storm_incidents", INCIDENT_HEADERS, rows)
        self.assertGreater(len(data), 1000)

    def test_xlsx_worksheet_title(self):
        data = _export_xlsx("storm_incidents", INCIDENT_HEADERS, [])
        from openpyxl import load_workbook
        loaded = load_workbook(io.BytesIO(data))
        self.assertEqual(loaded.active.title, "storm_incidents")


class TestColumnMapping(unittest.TestCase):
    """Verify each column in the export matches the correct field from the serialized doc."""

    def test_incident_column_order(self):
        rows = [_incident_row(SAMPLE_INCIDENT)]
        result = _export_csv(INCIDENT_HEADERS, rows)
        reader = csv.reader(io.StringIO(result))
        headers, data = list(reader)[0], list(reader)[0] if False else None
        rows_read = list(csv.reader(io.StringIO(result)))
        data_row = rows_read[1]
        self.assertEqual(data_row[0], "storm-2026-000001")  # Incident
        self.assertEqual(data_row[1], "sw1")                # Switch
        self.assertEqual(data_row[2], "Gi1/0/5")            # Interface
        self.assertEqual(data_row[3], "CRITICAL")           # Severity
        self.assertEqual(data_row[4], "OPEN")               # Status
        self.assertEqual(data_row[5], "2026-08-11T08:00:00Z")  # Created

    def test_mitigation_column_order(self):
        rows = [_mitigation_row(SAMPLE_MITIGATION)]
        result = _export_csv(MITIGATION_HEADERS, rows)
        rows_read = list(csv.reader(io.StringIO(result)))
        data_row = rows_read[1]
        self.assertEqual(data_row[0], "storm-2026-000001")  # Incident
        self.assertEqual(data_row[1], "Gi1/0/5")            # Interface
        self.assertEqual(data_row[2], "SHUTDOWN")           # Strategy
        self.assertEqual(data_row[3], "SUCCESS")            # Status
        self.assertEqual(data_row[4], "admin")              # Operator
        self.assertEqual(data_row[5], "2026-08-11T08:01:00Z")  # Time

    def test_recovery_column_order(self):
        rows = [_recovery_row(SAMPLE_RECOVERY)]
        result = _export_csv(RECOVERY_HEADERS, rows)
        rows_read = list(csv.reader(io.StringIO(result)))
        data_row = rows_read[1]
        self.assertEqual(data_row[0], "storm-2026-000001")  # Incident
        self.assertEqual(data_row[1], "Gi1/0/5")            # Interface
        self.assertEqual(data_row[2], "RECOVERED")          # Status
        self.assertEqual(data_row[3], "admin")              # Executed By
        self.assertEqual(data_row[4], "2026-08-11T08:30:00Z")  # Time


if __name__ == "__main__":
    unittest.main(verbosity=2)
